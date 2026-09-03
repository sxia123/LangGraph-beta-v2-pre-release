"""Spreadsheet parsing and deciphering module.

Extracts structured metadata, column metrics, summary statistics, and
clean Markdown tables from Excel (.xlsx, .xlsm, .xls) and CSV/TSV files.
"""

import csv
import io
import os
import re
from typing import Any, Dict, List, Optional, Tuple, Union


def decipher_spreadsheet(
    source: Union[str, bytes],
    filename: Optional[str] = None,
    row_limit: int = 150,
) -> Dict[str, Any]:
    """Deciphers an Excel workbook or CSV file into structured metrics and markdown tables."""
    resolved_path: Optional[str] = None
    raw_bytes: Optional[bytes] = None
    target_filename = filename or ""

    if isinstance(source, bytes):
        raw_bytes = source
    elif isinstance(source, str):
        source_str = source.strip()
        if source_str.startswith("data:"):
            try:
                import base64

                _header, encoded = source_str.split(",", 1)
                raw_bytes = base64.b64decode(encoded)
            except Exception:
                raw_bytes = None
        elif os.path.isfile(source_str):
            resolved_path = source_str
            if not target_filename:
                target_filename = os.path.basename(source_str)
        elif os.path.isfile(os.path.join(os.getcwd(), source_str)):
            resolved_path = os.path.join(os.getcwd(), source_str)
            if not target_filename:
                target_filename = os.path.basename(resolved_path)
        else:
            # Check if source_str is base64 string
            if len(source_str) > 32 and (len(source_str) % 4 == 0) and re.match(r"^[A-Za-z0-9+/=\r\n]+$", source_str):
                try:
                    import base64

                    raw_bytes = base64.b64decode(source_str)
                except Exception:
                    raw_bytes = None

    if not target_filename and resolved_path:
        target_filename = os.path.basename(resolved_path)
    fn_lower = (target_filename or "").lower()

    # 1. Parse Excel (.xlsx, .xlsm, .xltx)
    if fn_lower.endswith((".xlsx", ".xlsm", ".xltx", ".xls")) or (
        raw_bytes and (raw_bytes.startswith(b"PK\x03\x04") or raw_bytes.startswith(b"\xd0\xcf\x11\xe0"))
    ):
        return _parse_excel(resolved_path, raw_bytes, target_filename, row_limit)

    # 2. Parse CSV / TSV
    if fn_lower.endswith((".csv", ".tsv")) or (
        isinstance(source, str) and ("," in source or "\t" in source) and "\n" in source
    ):
        csv_text = None
        if raw_bytes:
            try:
                csv_text = raw_bytes.decode("utf-8")
            except Exception:
                csv_text = raw_bytes.decode("latin-1", errors="replace")
        elif resolved_path:
            try:
                with open(resolved_path, "r", encoding="utf-8", errors="replace") as f:
                    csv_text = f.read()
            except Exception:
                pass
        elif isinstance(source, str):
            csv_text = source

        if csv_text:
            return _parse_csv_text(csv_text, target_filename, row_limit)

    return {
        "ok": False,
        "filename": target_filename,
        "sheet_names": [],
        "summary": f"Unable to decipher spreadsheet format for '{target_filename}'.",
        "markdown_tables": "",
        "deciphered_context": f"[Notice: Unable to decipher spreadsheet '{target_filename}'. Unsupported format.]",
        "metrics": {},
    }


def _parse_excel(
    file_path: Optional[str],
    raw_bytes: Optional[bytes],
    filename: str,
    row_limit: int,
) -> Dict[str, Any]:
    try:
        import openpyxl

        wb = None
        if raw_bytes:
            wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)
        elif file_path:
            wb = openpyxl.load_workbook(file_path, data_only=True)

        if wb is None:
            return {
                "ok": False,
                "filename": filename,
                "sheet_names": [],
                "summary": f"Could not load Excel file '{filename}'.",
                "markdown_tables": "",
                "deciphered_context": f"[Excel file '{filename}' could not be loaded.]",
                "metrics": {},
            }

        sheet_names = wb.sheetnames
        markdown_sections: List[str] = []
        metrics: Dict[str, Any] = {}
        summary_lines: List[str] = [
            f"**Workbook**: `{filename or 'Spreadsheet'}`",
            f"**Total Sheets**: {len(sheet_names)} ({', '.join(sheet_names)})",
            "",
        ]

        for s_name in sheet_names:
            ws = wb[s_name]
            all_rows = list(ws.iter_rows(values_only=True))
            if not all_rows:
                continue

            header_raw = all_rows[0]
            header = [
                str(c if c is not None else f"Col_{idx+1}").replace("\n", " ").replace("|", "\\|")
                for idx, c in enumerate(header_raw)
            ]

            data_rows = [r for r in all_rows[1:] if any(c is not None for c in r)]
            col_stats = _compute_column_stats(header, data_rows)
            metrics[s_name] = {
                "rows_count": len(data_rows),
                "columns_count": len(header),
                "columns": header,
                "column_stats": col_stats,
            }

            summary_lines.append(
                f"- **Sheet '{s_name}'**: {len(data_rows)} data rows, {len(header)} columns."
            )
            if col_stats:
                stat_highlights = [
                    f"{col}: sum={stat['sum']:,.2f}" if "sum" in stat else col
                    for col, stat in list(col_stats.items())[:3]
                ]
                summary_lines.append(f"  *Key Metrics*: {'; '.join(stat_highlights)}")

            md_table = [
                f"### Sheet: {s_name}",
                f"*Dimensions: {len(data_rows)} rows × {len(header)} columns*",
                "",
                "| " + " | ".join(header) + " |",
                "| " + " | ".join(["---"] * max(1, len(header))) + " |",
            ]
            for row in data_rows[:row_limit]:
                cells = [
                    str(c if c is not None else "").replace("\n", " ").replace("|", "\\|")
                    for c in row
                ]
                if len(cells) < len(header):
                    cells.extend([""] * (len(header) - len(cells)))
                md_table.append("| " + " | ".join(cells[: len(header)]) + " |")

            if len(data_rows) > row_limit:
                md_table.append(
                    f"\n*... ({len(data_rows) - row_limit} additional rows omitted for token efficiency)*"
                )

            markdown_sections.append("\n".join(md_table))

        markdown_tables = "\n\n".join(markdown_sections)
        executive_summary = "\n".join(summary_lines)
        deciphered_context = (
            f"## Deciphered Spreadsheet Analysis: {filename or 'Workbook'}\n\n"
            f"{executive_summary}\n\n"
            f"### Structured Data Tables\n\n{markdown_tables}"
        )

        return {
            "ok": True,
            "filename": filename,
            "sheet_names": sheet_names,
            "summary": executive_summary,
            "markdown_tables": markdown_tables,
            "deciphered_context": deciphered_context,
            "metrics": metrics,
        }
    except Exception as err:
        return {
            "ok": False,
            "filename": filename,
            "sheet_names": [],
            "summary": f"Error parsing Excel spreadsheet '{filename}': {err}",
            "markdown_tables": "",
            "deciphered_context": f"[Notice: Unable to parse Excel file '{filename}' ({err}).]",
            "metrics": {},
        }


def _parse_csv_text(csv_text: str, filename: str, row_limit: int) -> Dict[str, Any]:
    try:
        sample = csv_text[:4096]
        delim = "\t" if "\t" in sample and sample.count("\t") > sample.count(",") else ","
        reader = list(csv.reader(io.StringIO(csv_text), delimiter=delim))
        if not reader:
            return {
                "ok": False,
                "filename": filename,
                "sheet_names": [],
                "summary": "Empty CSV/TSV content.",
                "markdown_tables": "",
                "deciphered_context": "[CSV content was empty.]",
                "metrics": {},
            }

        header = [str(c).replace("\n", " ").replace("|", "\\|") for c in reader[0]]
        data_rows = [r for r in reader[1:] if any(c.strip() for c in r)]
        col_stats = _compute_column_stats(header, data_rows)

        summary_lines = [
            f"**File**: `{filename or 'CSV Data'}`",
            f"**Rows**: {len(data_rows)}, **Columns**: {len(header)}",
        ]
        if col_stats:
            stat_highlights = [
                f"{col}: sum={stat['sum']:,.2f}" if "sum" in stat else col
                for col, stat in list(col_stats.items())[:3]
            ]
            summary_lines.append(f"**Key Metrics**: {'; '.join(stat_highlights)}")

        md_table = [
            f"### File: {filename or 'CSV Table'}",
            "",
            "| " + " | ".join(header) + " |",
            "| " + " | ".join(["---"] * max(1, len(header))) + " |",
        ]
        for row in data_rows[:row_limit]:
            cells = [str(c).replace("\n", " ").replace("|", "\\|") for c in row]
            if len(cells) < len(header):
                cells.extend([""] * (len(header) - len(cells)))
            md_table.append("| " + " | ".join(cells[: len(header)]) + " |")

        if len(data_rows) > row_limit:
            md_table.append(
                f"\n*... ({len(data_rows) - row_limit} additional rows omitted for token efficiency)*"
            )

        markdown_tables = "\n".join(md_table)
        executive_summary = "\n".join(summary_lines)
        deciphered_context = (
            f"## Deciphered Spreadsheet Analysis: {filename or 'CSV Data'}\n\n"
            f"{executive_summary}\n\n"
            f"### Structured Data Table\n\n{markdown_tables}"
        )

        return {
            "ok": True,
            "filename": filename,
            "sheet_names": ["Data"],
            "summary": executive_summary,
            "markdown_tables": markdown_tables,
            "deciphered_context": deciphered_context,
            "metrics": {
                "Data": {
                    "rows_count": len(data_rows),
                    "columns_count": len(header),
                    "columns": header,
                    "column_stats": col_stats,
                }
            },
        }
    except Exception as err:
        return {
            "ok": False,
            "filename": filename,
            "sheet_names": [],
            "summary": f"Error parsing CSV/TSV: {err}",
            "markdown_tables": "",
            "deciphered_context": f"[Notice: Unable to parse CSV ({err}).]",
            "metrics": {},
        }


def _compute_column_stats(
    header: List[str], data_rows: List[Union[List[Any], Tuple[Any, ...]]]
) -> Dict[str, Dict[str, float]]:
    stats: Dict[str, Dict[str, float]] = {}
    for idx, col_name in enumerate(header):
        vals: List[float] = []
        for row in data_rows:
            if idx < len(row):
                val = row[idx]
                if val is not None and not isinstance(val, bool):
                    if isinstance(val, (int, float)):
                        vals.append(float(val))
                    elif isinstance(val, str):
                        clean_v = val.replace("$", "").replace(",", "").replace("%", "").strip()
                        try:
                            vals.append(float(clean_v))
                        except ValueError:
                            pass
        if len(vals) >= max(2, len(data_rows) // 2):
            stats[col_name] = {
                "sum": sum(vals),
                "avg": sum(vals) / len(vals),
                "min": min(vals),
                "max": max(vals),
                "count": float(len(vals)),
            }
    return stats

