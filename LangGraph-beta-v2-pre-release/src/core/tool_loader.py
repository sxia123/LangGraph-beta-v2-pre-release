"""Tool loader: a central registry of imperative tools that pipeline agents can use.

Agents (graph nodes) get a `ToolLoader` instance and can:
  - call any registered tool by name:  loader.run("github_pull", repo_dir="...")
  - discover what's available:         loader.list_tools()
  - render a description block for a soul/system prompt: loader.prompt_block()
  - automatic checkpointing before any tool execution.
"""

import inspect
import os
import time
from typing import Any, Callable, Dict, List, Optional


class Tool:
    """A registered tool: a callable plus metadata for introspection."""

    def __init__(
        self,
        name: str,
        func: Callable,
        description: str = "",
        needs_llm: bool = False,
    ):
        self.name = name
        self.func = func
        doc_lines = (inspect.getdoc(func) or "").strip().splitlines()
        self.description = description or (doc_lines[0] if doc_lines else "")
        self.needs_llm = needs_llm

        params = [
            p.name
            for p in inspect.signature(func).parameters.values()
            if p.name != "self"
        ]
        self.params = params

    def to_dict(self) -> Dict[str, Any]:
        return {
            "name": self.name,
            "description": self.description,
            "params": self.params,
        }


class ToolLoader:
    """Registry + executor for imperative tools with automatic checkpointing."""

    def __init__(self, llm_client: Optional[Any] = None):
        self.llm_client = llm_client
        self._tools: Dict[str, Tool] = {}
        self._register_builtins()

    # ------------------------------------------------------------------ #
    # Registration
    # ------------------------------------------------------------------ #
    def register(
        self,
        name: str,
        func: Callable,
        description: str = "",
        needs_llm: bool = False,
    ) -> "ToolLoader":
        """Register (or overwrite) a tool. Returns self for chaining."""
        self._tools[name] = Tool(name, func, description=description, needs_llm=needs_llm)
        return self

    def _register_builtins(self) -> None:
        # 1. Web search (DuckDuckGo via LLM client or direct runner)
        self.register(
            "web_search",
            self._tool_web_search,
            description="Search the web via DuckDuckGo and return formatted markdown results.",
            needs_llm=True,
        )

        # 2. Wikipedia search
        self.register(
            "wikipedia",
            self._tool_wikipedia,
            description="Search Wikipedia for encyclopedic facts and summary pages.",
        )

        # 3. ArXiv scientific paper search
        self.register(
            "arxiv",
            self._tool_arxiv,
            description="Search ArXiv for academic publications, research abstracts, and papers.",
        )

        # 4. Python REPL execution
        self.register(
            "python_repl",
            self._tool_python_repl,
            description="Execute Python code in an interactive REPL environment and capture stdout/result.",
        )

        # 5. Math expression evaluation
        self.register(
            "math_eval",
            self._tool_math_eval,
            description="Evaluate high-performance mathematical expressions via NumExpr.",
        )

        # 6. Web scraping & HTML extraction
        self.register(
            "web_scrape",
            self._tool_web_scrape,
            description="Fetch and extract readable text content from a given URL via BeautifulSoup.",
        )

        # 7. Document conversion to Markdown (MarkItDown)
        self.register(
            "doc_convert",
            self._tool_doc_convert,
            description="Convert Office documents (DOCX, PPTX, XLSX), PDFs, or HTML into clean Markdown.",
        )

        # 8. Docling document parser
        self.register(
            "docling_parse",
            self._tool_docling_parse,
            description="Parse structured documents (PDF, DOCX, etc.) using IBM Docling parser.",
        )

        # 9. Vector memory search (ChromaDB)
        self.register(
            "vector_memory",
            self._tool_vector_memory,
            description="Query ChromaDB vector memory for semantically relevant memory chunks.",
        )

        # 10. File write (with automatic pre-file-change checkpointing)
        self.register(
            "file_write",
            self._tool_file_write,
            description="Write or overwrite a file at file_path with content (automatically creates pre-change checkpoint).",
        )

        # 11. File edit (with automatic pre-file-change checkpointing)
        self.register(
            "file_edit",
            self._tool_file_edit,
            description="Edit a file by finding and replacing a target substring (automatically creates pre-change checkpoint).",
        )

        # 12. GitHub tools
        try:
            from src.core import github_tool

            self.register("github_clone", github_tool.clone_repo,
                          description="Clone a git/GitHub repository into a local directory.")
            self.register("github_pull", github_tool.pull_repo,
                          description="Pull the latest changes for a local git repository.")
            self.register("github_push", github_tool.push_repo,
                          description="Push local commits to a git remote.")
            self.register("github_commit_push", github_tool.commit_and_push,
                          description="Stage all changes, commit, and push to the remote.")
            self.register("github_status", github_tool.get_repo_status,
                          description="Return a short git status summary (branch, ahead/behind, dirty files).")
        except Exception:
            pass

    # ------------------------------------------------------------------ #
    # Tool Implementations
    # ------------------------------------------------------------------ #
    def _tool_web_search(self, query: str, max_results: int = 5, max_chars: int = 8000) -> Dict[str, Any]:
        if self.llm_client is None:
            return {"ok": False, "message": "web_search requires an LLM client.", "returncode": None}
        try:
            raw = self.llm_client.search_web(query, max_results=max_results, max_tokens=12000)
            context = self.llm_client.build_search_context(raw, max_chars=max_chars)
            return {"ok": True, "message": context, "returncode": 0}
        except Exception as err:
            return {"ok": False, "message": f"web_search failed: {err}", "returncode": None}

    def _tool_wikipedia(self, query: str, sentences: int = 4) -> Dict[str, Any]:
        try:
            import wikipedia
            try:
                res = wikipedia.summary(query, sentences=sentences, auto_suggest=False)
                return {"ok": True, "message": res, "returncode": 0}
            except Exception:
                search_res = wikipedia.search(query, results=3)
                if search_res:
                    page = wikipedia.page(search_res[0], auto_suggest=False)
                    return {"ok": True, "message": page.summary[:1500], "returncode": 0}
                return {"ok": False, "message": f"No Wikipedia results found for '{query}'.", "returncode": None}
        except Exception as err:
            return {"ok": False, "message": f"Wikipedia query error: {err}", "returncode": None}

    def _tool_arxiv(self, query: str, max_results: int = 3) -> Dict[str, Any]:
        try:
            import arxiv
            client = arxiv.Client()
            search = arxiv.Search(query=query, max_results=max_results, sort_by=arxiv.SortCriterion.Relevance)
            papers = []
            for r in client.results(search):
                papers.append(
                    f"Title: {r.title}\nAuthors: {', '.join(a.name for a in r.authors)}\nPublished: {r.published.strftime('%Y-%m-%d')}\nSummary: {r.summary[:500]}...\nURL: {r.entry_id}"
                )
            if not papers:
                return {"ok": True, "message": "No arXiv papers found.", "returncode": 0}
            return {"ok": True, "message": "\n\n---\n\n".join(papers), "returncode": 0}
        except Exception as err:
            return {"ok": False, "message": f"ArXiv search error: {err}", "returncode": None}

    def _tool_python_repl(self, code: str, run_id: Optional[str] = None) -> Dict[str, Any]:
        try:
            # Auto-detect file write operations in code to checkpoint before execution
            import re

            from src.core.checkpointer import checkpoint_before_file_change

            file_matches = re.findall(r"""(?:open|Path)\s*\(\s*['"]([^'"]+)['"]\s*,\s*['"][wa\+]""", code)
            file_matches += re.findall(r"""Path\s*\(\s*['"]([^'"]+)['"]\s*\)\.write_""", code)
            for f_match in set(file_matches):
                try:
                    checkpoint_before_file_change(
                        f_match,
                        run_id=run_id,
                        metadata={"tool": "python_repl", "inferred_from_code": True},
                    )
                except Exception:
                    pass

            try:
                from langchain_experimental.utilities import PythonREPL

                repl = PythonREPL()
                res = repl.run(code)
                return {
                    "ok": True,
                    "message": res if res else "Code executed successfully with no stdout output.",
                    "returncode": 0,
                }
            except (ImportError, ModuleNotFoundError):
                import contextlib
                import io

                stdout_buf = io.StringIO()
                exec_globals: Dict[str, Any] = {"__name__": "__main__"}
                with contextlib.redirect_stdout(stdout_buf), contextlib.redirect_stderr(stdout_buf):
                    exec(code, exec_globals)
                out = stdout_buf.getvalue().strip()
                return {
                    "ok": True,
                    "message": out if out else "Code executed successfully with no stdout output.",
                    "returncode": 0,
                }
        except Exception as err:
            return {"ok": False, "message": f"Python REPL execution error: {err}", "returncode": None}

    def _tool_file_write(
        self, file_path: str, content: str, run_id: Optional[str] = None
    ) -> Dict[str, Any]:
        """Writes content to file_path with mandatory pre- and post-file-change checkpoints."""
        try:
            from src.core.checkpointer import (
                checkpoint_after_file_change,
                checkpoint_before_file_change,
            )

            abs_path = os.path.abspath(file_path)
            pre_cp = checkpoint_before_file_change(
                abs_path,
                run_id=run_id,
                metadata={"tool": "file_write", "action": "write"},
            )

            os.makedirs(os.path.dirname(abs_path), exist_ok=True)
            with open(abs_path, "w", encoding="utf-8") as f:
                f.write(content)

            post_cp = checkpoint_after_file_change(
                abs_path,
                pre_checkpoint_id=pre_cp["checkpoint_id"],
                run_id=run_id,
                metadata={"tool": "file_write"},
            )

            return {
                "ok": True,
                "message": f"File '{file_path}' written successfully ({len(content)} chars).",
                "checkpoint_id": pre_cp["checkpoint_id"],
                "post_checkpoint_id": post_cp["checkpoint_id"],
                "returncode": 0,
            }
        except Exception as err:
            return {"ok": False, "message": f"File write error: {err}", "returncode": None}

    def _tool_file_edit(
        self, file_path: str, target: str, replacement: str, run_id: Optional[str] = None
    ) -> Dict[str, Any]:
        """Edits an existing file by replacing target with replacement with mandatory checkpointing."""
        try:
            from src.core.checkpointer import (
                checkpoint_after_file_change,
                checkpoint_before_file_change,
            )

            abs_path = os.path.abspath(file_path)
            if not os.path.isfile(abs_path):
                return {"ok": False, "message": f"File '{file_path}' does not exist.", "returncode": None}

            pre_cp = checkpoint_before_file_change(
                abs_path,
                run_id=run_id,
                metadata={"tool": "file_edit", "action": "edit"},
            )

            with open(abs_path, "r", encoding="utf-8") as f:
                current_text = f.read()

            if target not in current_text:
                return {
                    "ok": False,
                    "message": f"Target string not found in '{file_path}'.",
                    "checkpoint_id": pre_cp["checkpoint_id"],
                    "returncode": None,
                }

            new_text = current_text.replace(target, replacement)
            with open(abs_path, "w", encoding="utf-8") as f:
                f.write(new_text)

            post_cp = checkpoint_after_file_change(
                abs_path,
                pre_checkpoint_id=pre_cp["checkpoint_id"],
                run_id=run_id,
                metadata={"tool": "file_edit"},
            )

            return {
                "ok": True,
                "message": f"File '{file_path}' edited successfully.",
                "checkpoint_id": pre_cp["checkpoint_id"],
                "post_checkpoint_id": post_cp["checkpoint_id"],
                "returncode": 0,
            }
        except Exception as err:
            return {"ok": False, "message": f"File edit error: {err}", "returncode": None}

    def _tool_math_eval(self, expression: str) -> Dict[str, Any]:
        try:
            try:
                import numexpr
                res = numexpr.evaluate(expression)
            except ImportError:
                import ast
                import operator
                ops = {
                    ast.Add: operator.add,
                    ast.Sub: operator.sub,
                    ast.Mult: operator.mul,
                    ast.Div: operator.truediv,
                    ast.FloorDiv: operator.floordiv,
                    ast.Mod: operator.mod,
                    ast.Pow: operator.pow,
                    ast.USub: operator.neg,
                    ast.UAdd: operator.pos,
                }

                def _eval_node(node: Any) -> Any:
                    if isinstance(node, ast.Constant) and isinstance(node.value, (int, float)):
                        return node.value
                    if isinstance(node, ast.BinOp):
                        left = _eval_node(node.left)
                        right = _eval_node(node.right)
                        return ops[type(node.op)](left, right)
                    if isinstance(node, ast.UnaryOp):
                        operand = _eval_node(node.operand)
                        return ops[type(node.op)](operand)
                    raise ValueError(f"Unsupported math expression '{expression}'")

                tree = ast.parse(expression.strip(), mode="eval")
                res = _eval_node(tree.body)

            return {"ok": True, "message": str(res), "returncode": 0}
        except Exception as err:
            return {"ok": False, "message": f"Math evaluation error: {err}", "returncode": None}


    def _tool_web_scrape(self, url: str, max_chars: int = 6000) -> Dict[str, Any]:
        try:
            import requests
            from bs4 import BeautifulSoup
            headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
            resp = requests.get(url, headers=headers, timeout=12)
            resp.raise_for_status()
            soup = BeautifulSoup(resp.text, "html.parser")
            for s in soup(["script", "style", "nav", "footer"]):
                s.extract()
            text = soup.get_text(separator="\n", strip=True)
            return {"ok": True, "message": text[:max_chars], "returncode": 0}
        except Exception as err:
            return {"ok": False, "message": f"Web scrape error: {err}", "returncode": None}

    def _tool_doc_convert(self, file_path: str) -> Dict[str, Any]:
        try:
            # 1. Native Excel spreadsheet parsing (.xlsx, .xls, .xlsm)
            if file_path.lower().endswith((".xlsx", ".xls", ".xlsm")):
                try:
                    import openpyxl

                    wb = openpyxl.load_workbook(file_path, data_only=True)
                    md_sheets = []
                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        rows = list(ws.iter_rows(values_only=True))
                        if not rows:
                            continue
                        header = [str(c if c is not None else "") for c in rows[0]]
                        md_table = [
                            f"### Sheet: {sheet_name}",
                            "",
                            "| " + " | ".join(header) + " |",
                            "| " + " | ".join(["---"] * len(header)) + " |",
                        ]
                        for row in rows[1:]:
                            if any(c is not None for c in row):
                                cells = [str(c if c is not None else "") for c in row]
                                md_table.append("| " + " | ".join(cells) + " |")
                        md_sheets.append("\n".join(md_table))
                    if md_sheets:
                        return {"ok": True, "message": "\n\n".join(md_sheets)[:16000], "returncode": 0}
                except Exception:
                    pass

            # 2. CSV / TSV spreadsheet parsing
            if file_path.lower().endswith((".csv", ".tsv")):
                try:
                    import csv

                    delim = "\t" if file_path.lower().endswith(".tsv") else ","
                    with open(file_path, "r", encoding="utf-8", errors="replace") as f:
                        reader = list(csv.reader(f, delimiter=delim))
                    if reader:
                        header = [str(c) for c in reader[0]]
                        md_table = [
                            "| " + " | ".join(header) + " |",
                            "| " + " | ".join(["---"] * len(header)) + " |",
                        ]
                        for row in reader[1:]:
                            md_table.append("| " + " | ".join(str(c) for c in row) + " |")
                        return {"ok": True, "message": "\n".join(md_table)[:16000], "returncode": 0}
                except Exception:
                    pass

            # 3. MarkItDown converter
            from markitdown import MarkItDown

            md = MarkItDown()
            res = md.convert(file_path)
            return {"ok": True, "message": res.text_content[:8000], "returncode": 0}
        except Exception as err:
            try:
                with open(file_path, "r", encoding="utf-8", errors="replace") as f:
                    return {"ok": True, "message": f.read()[:8000], "returncode": 0}
            except Exception:
                return {"ok": False, "message": f"Document conversion error: {err}", "returncode": None}

    def _tool_docling_parse(self, file_path: str) -> Dict[str, Any]:
        try:
            from docling.document_converter import DocumentConverter
            converter = DocumentConverter()
            res = converter.convert(file_path)
            md_text = res.document.export_to_markdown()
            return {"ok": True, "message": md_text[:8000], "returncode": 0}
        except Exception as err:
            return {"ok": False, "message": f"Docling parsing error: {err}", "returncode": None}

    def _tool_vector_memory(self, query: str, n_results: int = 5) -> Dict[str, Any]:
        try:
            import chromadb
            client = chromadb.Client()
            collection = client.get_or_create_collection("chart_pipeline_memory")
            res = collection.query(query_texts=[query], n_results=n_results)
            docs = res.get("documents", [[]])[0]
            return {"ok": True, "message": "\n".join(docs) if docs else "No vector memory matches found.", "returncode": 0}
        except Exception as err:
            return {"ok": False, "message": f"Vector memory query error: {err}", "returncode": None}

    # ------------------------------------------------------------------ #
    # ------------------------------------------------------------------ #
    # Execution with Pre- and Post-Tool Checkpointing
    # ------------------------------------------------------------------ #
    def _create_pre_checkpoint(
        self,
        tool_name: str,
        run_id: Optional[str] = None,
        kwargs: Optional[Dict[str, Any]] = None,
        metadata: Optional[Dict[str, Any]] = None,
    ) -> Optional[str]:
        """Creates a persistent checkpoint in SQLite before tool execution."""
        try:
            from src.core.memory_store import save_memory

            meta = {
                "checkpoint_type": "pre_tool_execution",
                "tool": tool_name,
            }
            if metadata:
                meta.update(metadata)

            checkpoint_entry = {
                "run_id": run_id,
                "event": f"checkpoint_before_{tool_name}",
                "input": kwargs or {},
                "result": f"Checkpoint created before executing tool [{tool_name}].",
                "metadata": meta,
                "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
            }
            return save_memory(checkpoint_entry)
        except Exception:
            return None

    def _create_checkpoint(
        self,
        tool_name: str,
        run_id: Optional[str] = None,
        kwargs: Optional[Dict[str, Any]] = None,
        metadata: Optional[Dict[str, Any]] = None,
    ) -> Optional[str]:
        """Backwards-compatible alias for creating a pre-tool checkpoint."""
        return self._create_pre_checkpoint(tool_name, run_id=run_id, kwargs=kwargs, metadata=metadata)

    def _create_post_checkpoint(
        self,
        tool_name: str,
        result: Dict[str, Any],
        duration_s: float,
        run_id: Optional[str] = None,
        kwargs: Optional[Dict[str, Any]] = None,
        pre_checkpoint_id: Optional[str] = None,
        metadata: Optional[Dict[str, Any]] = None,
    ) -> Optional[str]:
        """Creates a persistent checkpoint in SQLite after tool execution."""
        try:
            from src.core.memory_store import save_memory

            msg_val = result.get("message", "")
            summary_val = msg_val[:400] if isinstance(msg_val, str) else str(msg_val)[:400]

            meta = {
                "checkpoint_type": "post_tool_execution",
                "tool": tool_name,
                "ok": result.get("ok", False),
                "returncode": result.get("returncode"),
                "duration_s": round(duration_s, 4),
                "pre_checkpoint_id": pre_checkpoint_id,
            }
            if metadata:
                meta.update(metadata)

            checkpoint_entry = {
                "run_id": run_id,
                "event": f"checkpoint_after_{tool_name}",
                "input": kwargs or {},
                "result": summary_val,
                "metadata": meta,
                "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
            }
            return save_memory(checkpoint_entry)
        except Exception:
            return None

    def _create_error_checkpoint(
        self,
        tool_name: str,
        error_message: str,
        duration_s: float,
        run_id: Optional[str] = None,
        kwargs: Optional[Dict[str, Any]] = None,
        pre_checkpoint_id: Optional[str] = None,
        metadata: Optional[Dict[str, Any]] = None,
    ) -> Optional[str]:
        """Creates a persistent checkpoint in SQLite when tool execution encounters an error."""
        try:
            from src.core.memory_store import save_memory

            meta = {
                "checkpoint_type": "tool_error",
                "tool": tool_name,
                "ok": False,
                "duration_s": round(duration_s, 4),
                "pre_checkpoint_id": pre_checkpoint_id,
                "error": error_message,
            }
            if metadata:
                meta.update(metadata)

            checkpoint_entry = {
                "run_id": run_id,
                "event": f"checkpoint_error_{tool_name}",
                "input": kwargs or {},
                "result": error_message,
                "metadata": meta,
                "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
            }
            return save_memory(checkpoint_entry)
        except Exception:
            return None

    def run(
        self,
        name: str,
        run_id: Optional[str] = None,
        create_checkpoint: bool = True,
        metadata: Optional[Dict[str, Any]] = None,
        **kwargs: Any,
    ) -> Dict[str, Any]:
        """Execute a tool by name with automatic pre- and post-tool checkpointing."""
        tool = self._tools.get(name)
        if tool is None:
            return {
                "ok": False,
                "tool": name,
                "message": f"Unknown tool '{name}'. Available: {', '.join(self.list_tools())}",
                "returncode": None,
                "checkpoint_id": None,
                "post_checkpoint_id": None,
            }

        # 1. Pre-tool checkpoint
        checkpoint_id = None
        if create_checkpoint:
            checkpoint_id = self._create_pre_checkpoint(name, run_id=run_id, kwargs=kwargs, metadata=metadata)

        start_time = time.time()

        # 2. Tool Execution
        try:
            call_kwargs = dict(kwargs)
            try:
                import inspect

                sig = inspect.signature(tool.func)
                if "run_id" in sig.parameters and "run_id" not in call_kwargs:
                    call_kwargs["run_id"] = run_id
            except Exception:
                pass

            if tool.needs_llm:
                result = (
                    tool.func(self.llm_client, **call_kwargs)
                    if _first_param_is_client(tool.func)
                    else tool.func(**call_kwargs)
                )
            else:
                result = tool.func(**call_kwargs)
        except TypeError as err:
            duration_s = time.time() - start_time
            err_msg = f"Invalid arguments for '{name}': {err}"
            post_id = None
            if create_checkpoint:
                post_id = self._create_error_checkpoint(
                    name,
                    error_message=err_msg,
                    duration_s=duration_s,
                    run_id=run_id,
                    kwargs=kwargs,
                    pre_checkpoint_id=checkpoint_id,
                    metadata=metadata,
                )
            return {
                "ok": False,
                "tool": name,
                "checkpoint_id": checkpoint_id,
                "post_checkpoint_id": post_id,
                "message": err_msg,
                "returncode": None,
            }
        except Exception as err:
            duration_s = time.time() - start_time
            err_msg = f"Tool '{name}' raised: {err}"
            post_id = None
            if create_checkpoint:
                post_id = self._create_error_checkpoint(
                    name,
                    error_message=err_msg,
                    duration_s=duration_s,
                    run_id=run_id,
                    kwargs=kwargs,
                    pre_checkpoint_id=checkpoint_id,
                    metadata=metadata,
                )
            return {
                "ok": False,
                "tool": name,
                "checkpoint_id": checkpoint_id,
                "post_checkpoint_id": post_id,
                "message": err_msg,
                "returncode": None,
            }

        duration_s = time.time() - start_time

        # Normalize result
        if isinstance(result, str):
            result = {"ok": True, "message": result, "returncode": 0}
        if not isinstance(result, dict):
            result = {"ok": True, "message": str(result), "returncode": 0}

        result.setdefault("tool", name)

        # 3. Post-tool checkpoint
        post_checkpoint_id = None
        if create_checkpoint:
            post_checkpoint_id = self._create_post_checkpoint(
                name,
                result=result,
                duration_s=duration_s,
                run_id=run_id,
                kwargs=kwargs,
                pre_checkpoint_id=checkpoint_id,
                metadata=metadata,
            )

        result["checkpoint_id"] = checkpoint_id
        result["post_checkpoint_id"] = post_checkpoint_id
        return result

    # ------------------------------------------------------------------ #
    # Introspection
    # ------------------------------------------------------------------ #
    def list_tools(self) -> List[str]:
        return sorted(self._tools.keys())

    def get(self, name: str) -> Optional[Tool]:
        return self._tools.get(name)

    def describe(self) -> List[Dict[str, Any]]:
        return [t.to_dict() for t in self._tools.values()]

    def prompt_block(self, max_chars: int = 2000) -> str:
        """Render a compact description of all tools for injection into a prompt."""
        if not self._tools:
            return "No tools available."
        lines = []
        for tool in self._tools.values():
            params = ", ".join(tool.params) if tool.params else "none"
            lines.append(f"- {tool.name}({params}): {tool.description}")
        block = "\n".join(lines)
        if len(block) > max_chars:
            block = block[:max_chars].rstrip() + "\n[...truncated...]"
        return block


def _first_param_is_client(func: Callable) -> bool:
    """Heuristic: does the tool's first parameter expect the LLM client?"""
    try:
        params = list(inspect.signature(func).parameters.values())
    except (TypeError, ValueError):
        return False
    if not params:
        return False
    return params[0].name in {"llm_client", "client", "llm"}


# ---------------------------------------------------------------------- #
# Module-level singleton so nodes can share one loader
# ---------------------------------------------------------------------- #
_default_loader: Optional[ToolLoader] = None


def get_tool_loader(llm_client: Optional[Any] = None) -> ToolLoader:
    """Return a shared ToolLoader, binding the given LLM client if provided."""
    global _default_loader
    if llm_client is not None or _default_loader is None:
        _default_loader = ToolLoader(llm_client=llm_client)
    elif _default_loader.llm_client is None:
        _default_loader.llm_client = llm_client
    return _default_loader


def checkpoint_tool(
    name: Optional[str] = None,
    run_id: Optional[str] = None,
    metadata: Optional[Dict[str, Any]] = None,
) -> Callable:
    """Decorator to wrap any standalone function with automated pre- and post-checkpoints."""
    def decorator(func: Callable) -> Callable:
        tool_name = name or func.__name__

        def wrapper(*args: Any, **kwargs: Any) -> Any:
            loader = get_tool_loader()
            all_kwargs = dict(kwargs)
            if args:
                sig = inspect.signature(func)
                param_names = list(sig.parameters.keys())
                for idx, arg_val in enumerate(args):
                    if idx < len(param_names):
                        all_kwargs[param_names[idx]] = arg_val

            pre_id = loader._create_pre_checkpoint(
                tool_name, run_id=run_id, kwargs=all_kwargs, metadata=metadata
            )
            start_t = time.time()
            try:
                res = func(*args, **kwargs)
                dur = time.time() - start_t
                res_dict = res if isinstance(res, dict) else {"ok": True, "message": str(res), "returncode": 0}
                loader._create_post_checkpoint(
                    tool_name,
                    result=res_dict,
                    duration_s=dur,
                    run_id=run_id,
                    kwargs=all_kwargs,
                    pre_checkpoint_id=pre_id,
                    metadata=metadata,
                )
                return res
            except Exception as err:
                dur = time.time() - start_t
                loader._create_error_checkpoint(
                    tool_name,
                    error_message=str(err),
                    duration_s=dur,
                    run_id=run_id,
                    kwargs=all_kwargs,
                    pre_checkpoint_id=pre_id,
                    metadata=metadata,
                )
                raise

        return wrapper

    return decorator

