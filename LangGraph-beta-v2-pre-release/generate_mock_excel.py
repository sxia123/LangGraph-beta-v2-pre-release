"""Generates a rich, multi-sheet mock Excel workbook for enterprise financial analysis."""

import os

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def create_mock_excel(output_path: str = "mock_enterprise_financial_q1_2026.xlsx") -> str:
    wb = openpyxl.Workbook()

    # Style definitions
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    Font(name="Calibri", size=11, bold=True, color="1F4E79")
    total_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    total_font = Font(name="Calibri", size=11, bold=True)
    regular_font = Font(name="Calibri", size=11)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # ---------------------------------------------------------
    # Sheet 1: Revenue & ARR
    # ---------------------------------------------------------
    ws_rev = wb.active
    ws_rev.title = "Revenue & ARR"
    ws_rev.views.sheetView[0].showGridLines = True

    rev_headers = [
        "Region", "Business Unit", "Product Line", "Q1 Target ($)",
        "Q1 Actual ($)", "Variance ($)", "YoY Growth %", "Status", "Strategic Notes"
    ]
    ws_rev.append(rev_headers)

    rev_rows = [
        ["Americas", "Enterprise AI", "Agentic Orchestration Platform", 3200000, 3850000, 650000, 0.425, "Exceeded", "High tier Fortune 500 expansions"],
        ["Americas", "Cloud Infra", "Dedicated GPU Pods (H100/A100)", 2400000, 2750000, 350000, 0.280, "Exceeded", "Strong AI fine-tuning compute demand"],
        ["Americas", "Developer Tools", "Copilot IDE & CLI Subscriptions", 1100000, 1080000, -20000, 0.120, "On Track", "Transitioning to per-seat enterprise tiers"],
        ["EMEA", "Enterprise AI", "Agentic Orchestration Platform", 1800000, 1950000, 150000, 0.310, "Exceeded", "DAX and FTSE 100 enterprise pilots closing"],
        ["EMEA", "Cloud Infra", "European Sovereign Cloud Clusters", 1500000, 1420000, -80000, 0.085, "On Track", "GDPR compliance certifications finalized"],
        ["EMEA", "Developer Tools", "Copilot IDE & CLI Subscriptions", 850000, 920000, 70000, 0.220, "Exceeded", "Strong developer community adoption in UK/DE"],
        ["APAC", "Enterprise AI", "Agentic Orchestration Platform", 1400000, 1820000, 420000, 0.580, "Exceeded", "Explosive growth in Singapore & Tokyo fintech"],
        ["APAC", "Cloud Infra", "Dedicated GPU Pods (H100/A100)", 950000, 1100000, 150000, 0.340, "Exceeded", "Cross-regional low-latency inference routing"],
        ["APAC", "Professional Services", "Agentic Architecture Consulting", 500000, 480000, -20000, -0.040, "On Track", "Partner enablement reducing custom consulting"],
        ["LATAM", "Enterprise AI", "Agentic Orchestration Platform", 550000, 640000, 90000, 0.380, "Exceeded", "São Paulo and Mexico City banking hubs"],
        ["LATAM", "Developer Tools", "Copilot IDE & CLI Subscriptions", 350000, 390000, 40000, 0.250, "Exceeded", "Tech hub startup adoption"],
    ]

    for r in rev_rows:
        ws_rev.append(r)

    # Total row
    total_target = sum(r[3] for r in rev_rows)
    total_actual = sum(r[4] for r in rev_rows)
    total_variance = total_actual - total_target
    total_growth = (total_actual - total_target) / total_target
    ws_rev.append(["Total / Summary", "All Units", "Consolidated", total_target, total_actual, total_variance, total_growth, "Exceeded (+11.8%)", "Overall revenue ahead of annual plan"])

    # ---------------------------------------------------------
    # Sheet 2: Operating Expenses
    # ---------------------------------------------------------
    ws_exp = wb.create_sheet(title="Operating Expenses")
    ws_exp.views.sheetView[0].showGridLines = True

    exp_headers = ["Cost Center", "Sub-Category", "Q1 Budget ($)", "Q1 Actual ($)", "Variance ($)", "Burn Rate %", "Status", "Mitigation & Action Plan"]
    ws_exp.append(exp_headers)

    exp_rows = [
        ["R&D - Engineering", "AI Model Training & Research", 2800000, 2650000, -150000, 0.946, "Under Budget", "Quantization & FlashAttention-3 efficiencies"],
        ["R&D - Engineering", "Core Infrastructure & Platform", 1900000, 1850000, -50000, 0.974, "On Budget", "Distributed state checkpointer optimizations"],
        ["R&D - Engineering", "Product Design & UX", 600000, 580000, -20000, 0.967, "On Budget", "Standardized design system component rollout"],
        ["Infrastructure / Hosting", "Data Center Co-location & GPU Power", 1400000, 1550000, 150000, 1.107, "Over Budget", "Surge in multi-agent graph execution load"],
        ["Sales & Marketing", "Enterprise Field Sales & SDRs", 1600000, 1520000, -80000, 0.950, "Under Budget", "Hiring timeline shifted slightly into Q2"],
        ["Sales & Marketing", "Global Events & Developer Conferences", 750000, 790000, 40000, 1.053, "Over Budget", "Sponsorship for AI World Summit in SF"],
        ["Customer Success", "Enterprise Solutions Architecture", 900000, 880000, -20000, 0.978, "On Budget", "Automated customer diagnostic agents deployed"],
        ["General & Administrative", "Legal, Compliance & Patents", 500000, 560000, 60000, 1.120, "Over Budget", "EU AI Act regulatory compliance auditing"],
        ["General & Administrative", "People Ops, Facilities & IT", 650000, 630000, -20000, 0.969, "On Budget", "Consolidated SaaS subscription redundancy"],
    ]

    for r in exp_rows:
        ws_exp.append(r)

    total_exp_budget = sum(r[2] for r in exp_rows)
    total_exp_actual = sum(r[3] for r in exp_rows)
    total_exp_variance = total_exp_actual - total_exp_budget
    total_exp_burn = total_exp_actual / total_exp_budget
    ws_exp.append(["Total OPEX", "Consolidated", total_exp_budget, total_exp_actual, total_exp_variance, total_exp_burn, "Controlled", "Total spend strictly within 99.1% of forecast"])

    # ---------------------------------------------------------
    # Sheet 3: Executive KPIs
    # ---------------------------------------------------------
    ws_kpi = wb.create_sheet(title="Executive KPIs")
    ws_kpi.views.sheetView[0].showGridLines = True

    kpi_headers = ["Strategic Metric", "Category", "Target Q1 2026", "Actual Q1 2026", "YoY Change", "Industry Benchmark", "Executive Rating"]
    ws_kpi.append(kpi_headers)

    kpi_rows = [
        ["Annual Recurring Revenue (ARR) ($)", "Financial", 52000000, 58400000, 0.385, 45000000, "Elite (Top Decile)"],
        ["Net Revenue Retention (NRR) %", "Retention", 1.250, 1.340, 0.090, 1.180, "Exceptional (Strong Land & Expand)"],
        ["Gross Margin %", "Profitability", 0.780, 0.815, 0.035, 0.720, "Expanded (+350 bps)"],
        ["Rule of 40 Score", "Efficiency", 0.450, 0.535, 0.085, 0.400, "Exceeded (38.5% Growth + 15% Margin)"],
        ["Customer Acquisition Cost (CAC) Payback (Mo)", "Sales Efficiency", 12.0, 9.4, -2.6, 14.0, "Highly Accelerated"],
        ["Active Enterprise Customers (> $100k ARR)", "Scale", 450, 528, 0.280, 380, "78 Net New Additions in Q1"],
        ["Average Graph Workflow Completion Time (ms)", "Platform Performance", 450, 310, -0.311, 600, "31% Latency Reduction via Local Agents"],
    ]

    for r in kpi_rows:
        ws_kpi.append(r)

    # Format all sheets with beautiful header and cell styles
    for ws in [ws_rev, ws_exp, ws_kpi]:
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Style data rows
        max_row = ws.max_row
        max_col = ws.max_column
        for row_idx in range(2, max_row + 1):
            is_last = (row_idx == max_row and ws != ws_kpi)
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                if is_last:
                    cell.fill = total_fill
                    cell.font = total_font
                else:
                    cell.font = regular_font

                # Number formatting
                val = cell.value
                if isinstance(val, float):
                    if 0.0 <= abs(val) <= 2.0 and "Payback" not in str(ws.cell(row=row_idx, column=1).value):
                        cell.number_format = "0.0%"
                    else:
                        cell.number_format = "$#,##0"
                elif isinstance(val, int) and val > 1000:
                    cell.number_format = "$#,##0"

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or "")
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

    abs_output = os.path.abspath(output_path)
    wb.save(abs_output)
    print(f"Successfully generated mock Excel file at: {abs_output}")
    return abs_output


if __name__ == "__main__":
    create_mock_excel()
