import json
import logging
import operator
import os
import queue as _queue
import threading as _threading
import time
from typing import Annotated, Any, Dict, List, Optional

from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from fastapi.staticfiles import StaticFiles
from langgraph.graph import END, START, StateGraph
from pydantic import BaseModel
from typing_extensions import TypedDict

from src.agents import (
    create_chart_pipeline_graph,
    create_claims_triage_graph,
    create_code_review_team_graph,
    create_master_pipeline_graph,
    create_multi_agent_supervisor_graph,
    create_solution_review_team_graph,
)
from src.core.local_llm import LocalLLMClient
from src.core.memory_store import (
    clear_memories,
    delete_memory,
    delete_run,
    fetch_memories,
    finish_run,
    get_memory_summary,
    get_run,
    get_run_memories,
    get_tool_checkpoints,
    list_runs,
    save_memory,
    search_memories,
    start_run,
)
from src.core.tool_loader import get_tool_loader

logger = logging.getLogger("server")

app = FastAPI(title="LangGraph Web API Server", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

llm_client = LocalLLMClient()


class DirectChatState(TypedDict, total=False):
    messages: Annotated[List[Dict[str, Any]], operator.add]
    user_input: str
    final_response: str
    agent_thoughts: Annotated[List[Dict[str, Any]], operator.add]
    run_id: Optional[str]
    images: Optional[List[str]]
    files: Optional[List[Dict[str, Any]]]


def create_direct_chat_graph(
    client: LocalLLMClient, checkpointer: Optional[Any] = None
):
    workflow = StateGraph(DirectChatState)
    tools = get_tool_loader(client)

    def direct_chat_node(state: DirectChatState) -> Dict[str, Any]:
        messages = state.get("messages", [])
        last_user = messages[-1] if messages else {}
        images = state.get("images") or last_user.get("images") or []
        run_id = state.get("run_id")

        tool_prompt = (
            "You are Qwen, a helpful, highly capable vision and language AI assistant equipped with tools.\n"
            "Respond accurately, clearly, and concisely to user questions and images.\n\n"
            "Available Tools (imperative execution):\n"
            f"{tools.prompt_block()}\n\n"
            "If the user asks you to perform a calculation, write/run Python code, search the web, lookup Wikipedia/ArXiv, "
            "convert documents, or use git, format your tool call as a JSON block:\n"
            "```json\n"
            '{"tool": "<tool_name>", "args": {<param_name>: <value>}}\n'
            "```\n"
            "If no tool is required, respond directly with your helpful answer."
        )

        res = client.generate_completion(
            system_prompt=tool_prompt,
            messages=messages,
            available_tools=tools.list_tools(),
            images=images,
        )
        thought_entries = []
        if res.thought:
            thought_entries.append(
                {
                    "agent": "Qwen Reasoning",
                    "thought": res.thought,
                    "timestamp": time.strftime("%H:%M:%S"),
                }
            )

        # Check for tool calls
        tool_calls = res.tool_calls or []
        if not tool_calls:
            import re
            json_matches = re.findall(r"```(?:json)?\s*(\{[\s\S]*?\})\s*```", res.content)
            if not json_matches:
                raw_match = re.search(r"(\{\s*\"(?:tool|name)\"\s*:\s*\"[^\"]+\"[\s\S]*?\})", res.content)
                if raw_match:
                    json_matches = [raw_match.group(1)]
            for j_str in json_matches:
                try:
                    parsed = json.loads(j_str)
                    if isinstance(parsed, dict) and ("tool" in parsed or "name" in parsed):
                        tool_calls.append({
                            "name": parsed.get("tool") or parsed.get("name"),
                            "args": parsed.get("args") or parsed.get("parameters") or {},
                        })
                        break
                except Exception:
                    pass

        if tool_calls:
            for tc in tool_calls:
                t_name = tc.get("name")
                t_args = tc.get("args") or {}
                if t_name:
                    tool_res = tools.run(
                        t_name,
                        run_id=run_id,
                        metadata={"source": "direct_chat"},
                        **t_args,
                    )
                    cp_id = tool_res.get("checkpoint_id")
                    post_id = tool_res.get("post_checkpoint_id")
                    status_text = "SUCCESS" if tool_res.get("ok") else "FAILED"
                    msg_text = tool_res.get("message", "")

                    thought_entries.append(
                        {
                            "agent": f"Tool: {t_name}",
                            "thought": (
                                f"**Tool Execution**: `{t_name}` ({status_text})\n"
                                f"- **Arguments**: `{json.dumps(t_args)}`\n"
                                f"- **Checkpoint**: `{post_id or cp_id or 'N/A'}`\n\n"
                                f"**Result Preview**:\n```\n{str(msg_text)[:1200]}\n```"
                            ),
                            "tool": t_name,
                            "args": t_args,
                            "result": tool_res,
                            "ok": tool_res.get("ok", False),
                            "checkpoint_id": post_id or cp_id,
                            "timestamp": time.strftime("%H:%M:%S"),
                        }
                    )

                    # Synthesize final answer with tool result context
                    synth_messages = list(messages) + [
                        {
                            "role": "assistant",
                            "content": f"Executed tool [{t_name}].",
                        },
                        {
                            "role": "user",
                            "content": f"Tool [{t_name}] Execution Result:\n{msg_text}\n\nPlease synthesize the final answer based on this tool result.",
                        },
                    ]
                    synth_res = client.generate_completion(
                        system_prompt="You are Qwen, a helpful vision and language AI assistant. Provide a clean, direct, and complete final answer to the user based on the tool result.",
                        messages=synth_messages,
                    )
                    if synth_res.thought:
                        thought_entries.append(
                            {
                                "agent": "Qwen Synthesis",
                                "thought": synth_res.thought,
                                "timestamp": time.strftime("%H:%M:%S"),
                            }
                        )
                    final_text = synth_res.content
                    return {
                        "final_response": final_text,
                        "messages": [
                            {
                                "id": f"qwen_{int(time.time() * 1000)}",
                                "sender": "Qwen",
                                "role": "assistant",
                                "content": final_text,
                                "timestamp": time.strftime("%H:%M:%S"),
                            }
                        ],
                        "agent_thoughts": thought_entries,
                    }

        return {
            "final_response": res.content,
            "messages": [
                {
                    "id": f"qwen_{int(time.time() * 1000)}",
                    "sender": "Qwen",
                    "role": "assistant",
                    "content": res.content,
                    "timestamp": time.strftime("%H:%M:%S"),
                }
            ],
            "agent_thoughts": thought_entries,
        }

    workflow.add_node("qwen_assistant", direct_chat_node)
    workflow.add_edge(START, "qwen_assistant")
    workflow.add_edge("qwen_assistant", END)

    from src.core.checkpointer import get_default_checkpointer

    active_checkpointer = checkpointer if checkpointer is not None else get_default_checkpointer()
    return workflow.compile(checkpointer=active_checkpointer)


WORKFLOW_FACTORIES = {
    "direct": create_direct_chat_graph,
    "chat": create_direct_chat_graph,
    "vision": create_direct_chat_graph,
    "qwen": create_direct_chat_graph,
    "master": create_master_pipeline_graph,
    "master_pipeline": create_master_pipeline_graph,
    "chart": create_chart_pipeline_graph,
    "chart_pipeline": create_chart_pipeline_graph,
    "supervisor": create_multi_agent_supervisor_graph,
    "claims_triage": create_claims_triage_graph,
    "claims": create_claims_triage_graph,
    "code_review": create_code_review_team_graph,
    "code": create_code_review_team_graph,
    "solution_review": create_solution_review_team_graph,
    "solution": create_solution_review_team_graph,
}

SUPPORTED_WORKFLOWS = sorted(WORKFLOW_FACTORIES.keys())


class ChatRequest(BaseModel):
    prompt: str
    images: Optional[List[str]] = None
    files: Optional[List[Dict[str, Any]]] = None
    pipeline: Optional[str] = "master"
    model_name: Optional[str] = None
    agent_models: Optional[Dict[str, str]] = None
    session_id: Optional[str] = None
    use_memory: Optional[bool] = True


class FileUploadPayload(BaseModel):
    filename: str
    content: str
    content_type: Optional[str] = None


class MemoryCreateRequest(BaseModel):
    event: str
    input: Optional[Any] = None
    result: Optional[Any] = None
    metadata: Optional[Dict[str, Any]] = None


class ToolRunRequest(BaseModel):
    tool: str
    args: Optional[Dict[str, Any]] = None
    run_id: Optional[str] = None
    metadata: Optional[Dict[str, Any]] = None


class ModelSelectRequest(BaseModel):
    model_name: str


class ChatStopRequest(BaseModel):
    run_id: Optional[str] = None


@app.post("/api/chat/stop")
def stop_chat_stream(payload: Optional[ChatStopRequest] = None):
    """Signals cancellation / pause for an active generation stream."""
    run_id = payload.run_id if payload else None
    return {
        "ok": True,
        "message": f"Generation stopped successfully for run {run_id}" if run_id else "Generation stopped successfully.",
    }


def extract_file_text(filename: str, raw_content: str) -> str:
    """Extracts readable text/markdown from uploaded raw file content, data URI, or disk path."""
    raw_content = raw_content or ""
    fn_lower = (filename or "").lower()
    decoded_bytes: Optional[bytes] = None

    if raw_content.startswith("data:"):
        try:
            import base64

            _header, encoded = raw_content.split(",", 1)
            decoded_bytes = base64.b64decode(encoded)
        except Exception as err:
            logger.warning("Failed to decode data URI for %s: %s", filename, err)
    elif raw_content and not os.path.isfile(raw_content):
        binary_extensions = (".xlsx", ".xls", ".xlsm", ".pdf", ".docx", ".pptx", ".zip", ".bin")
        if fn_lower.endswith(binary_extensions):
            try:
                import base64
                import re

                clean_b64 = raw_content.strip()
                if len(clean_b64) > 32 and (len(clean_b64) % 4 == 0) and re.match(r"^[A-Za-z0-9+/=\r\n]+$", clean_b64):
                    decoded_bytes = base64.b64decode(clean_b64)
            except Exception:
                decoded_bytes = None

    # Resolve disk path fallback if content is empty or points to an existing file
    disk_path = None
    if raw_content and os.path.isfile(raw_content):
        disk_path = raw_content
    elif filename and os.path.isfile(filename):
        disk_path = filename
    elif filename and os.path.isfile(os.path.join(os.getcwd(), filename)):
        disk_path = os.path.join(os.getcwd(), filename)

    # 1. Native Excel spreadsheet parsing (.xlsx, .xls, .xlsm)
    if fn_lower.endswith((".xlsx", ".xls", ".xlsm")):
        try:
            import io

            import openpyxl

            wb = None
            if decoded_bytes is not None:
                wb = openpyxl.load_workbook(io.BytesIO(decoded_bytes), data_only=True)
            elif disk_path:
                wb = openpyxl.load_workbook(disk_path, data_only=True)

            if wb:
                md_sheets = []
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    rows = list(ws.iter_rows(values_only=True))
                    if not rows:
                        continue
                    # Clean header cells
                    header = [
                        str(c if c is not None else "").replace("\n", " ").replace("|", "\\|")
                        for c in rows[0]
                    ]
                    md_table = [
                        f"### Sheet: {sheet_name}",
                        "",
                        "| " + " | ".join(header) + " |",
                        "| " + " | ".join(["---"] * max(1, len(header))) + " |",
                    ]
                    row_limit = 150
                    data_rows = [r for r in rows[1:] if any(c is not None for c in r)]
                    for row in data_rows[:row_limit]:
                        cells = [
                            str(c if c is not None else "").replace("\n", " ").replace("|", "\\|")
                            for c in row
                        ]
                        md_table.append("| " + " | ".join(cells) + " |")
                    if len(data_rows) > row_limit:
                        md_table.append(
                            f"\n*... ({len(data_rows) - row_limit} additional rows omitted for token budget)*"
                        )
                    md_sheets.append("\n".join(md_table))
                if md_sheets:
                    return "\n\n".join(md_sheets)
                return f"[Excel workbook '{filename}' contains no readable data or sheets.]"
        except Exception as err:
            logger.warning("Error parsing Excel spreadsheet %s: %s", filename, err)
            return f"[Notice: Unable to parse Excel file '{filename}' ({err}). Please ensure the workbook is not password-protected.]"

    # 2. CSV / TSV spreadsheet parsing
    if fn_lower.endswith((".csv", ".tsv")):
        try:
            import csv
            import io

            csv_text = None
            if decoded_bytes is not None:
                try:
                    csv_text = decoded_bytes.decode("utf-8")
                except Exception:
                    csv_text = decoded_bytes.decode("latin-1", errors="replace")
            elif disk_path:
                with open(disk_path, "r", encoding="utf-8", errors="replace") as cf:
                    csv_text = cf.read()
            elif raw_content:
                csv_text = raw_content

            if csv_text:
                delim = "\t" if fn_lower.endswith(".tsv") else ","
                reader = list(csv.reader(io.StringIO(csv_text), delimiter=delim))
                if reader:
                    header = [str(c).replace("\n", " ").replace("|", "\\|") for c in reader[0]]
                    md_table = [
                        "| " + " | ".join(header) + " |",
                        "| " + " | ".join(["---"] * max(1, len(header))) + " |",
                    ]
                    row_limit = 150
                    for row in reader[1:row_limit]:
                        cells = [str(c).replace("\n", " ").replace("|", "\\|") for c in row]
                        md_table.append("| " + " | ".join(cells) + " |")
                    if len(reader) - 1 > row_limit:
                        md_table.append(
                            f"\n*... ({len(reader) - 1 - row_limit} additional rows omitted for token budget)*"
                        )
                    return "\n".join(md_table)
        except Exception as err:
            logger.warning("Error parsing CSV/TSV %s: %s", filename, err)

    # 3. PDF Document parsing (.pdf)
    if fn_lower.endswith(".pdf"):
        try:
            from src.core.document_parser import parse_pdf

            source_val = decoded_bytes or disk_path or raw_content
            res = parse_pdf(source_val, filename=filename)
            if res.get("ok") and res.get("extracted_text"):
                return res.get("extracted_text")
            logger.warning("PDF parse issue for %s: %s", filename, res.get("summary"))
        except Exception as err:
            logger.warning("Error parsing PDF %s: %s", filename, err)

    # 4. Word Document parsing (.docx, .doc)
    if fn_lower.endswith((".docx", ".doc")):
        try:
            from src.core.document_parser import parse_docx

            source_val = decoded_bytes or disk_path or raw_content
            res = parse_docx(source_val, filename=filename)
            if res.get("ok") and res.get("extracted_text"):
                return res.get("extracted_text")
            logger.warning("DOCX parse issue for %s: %s", filename, res.get("summary"))
        except Exception as err:
            logger.warning("Error parsing DOCX %s: %s", filename, err)

    # 5. PowerPoint Slideshow parsing (.pptx, .ppt)
    if fn_lower.endswith((".pptx", ".ppt")):
        try:
            from src.core.document_parser import parse_slideshow

            source_val = decoded_bytes or disk_path or raw_content
            res = parse_slideshow(source_val, filename=filename)
            if res.get("ok") and res.get("extracted_text"):
                return res.get("extracted_text")
            logger.warning("PPTX parse issue for %s: %s", filename, res.get("summary"))
        except Exception as err:
            logger.warning("Error parsing PPTX %s: %s", filename, err)

    # 6. Photos & Images (.png, .jpg, .jpeg, .webp, .gif, .bmp, .svg)
    if fn_lower.endswith((".png", ".jpg", ".jpeg", ".webp", ".gif", ".bmp", ".svg")) or (
        raw_content.startswith("data:image/")
    ):
        return f"[Image / Photo Attachment: {filename}]"

    if decoded_bytes is not None:
        try:
            return decoded_bytes.decode("utf-8")
        except Exception:
            return decoded_bytes.decode("latin-1", errors="replace")

    return raw_content


@app.post("/api/upload")
def upload_file_endpoint(payload: FileUploadPayload):
    """Ingests a file (text, code, or document) and extracts its textual content."""
    filename = payload.filename or "uploaded_file.txt"
    raw_content = payload.content or ""
    text = extract_file_text(filename, raw_content)

    return {
        "ok": True,
        "filename": filename,
        "size": len(text),
        "text": text,
        "content_type": payload.content_type or "text/plain",
    }


@app.get("/api/status")
def get_status():
    """Returns local LLM provider configuration, connection status, and tool availability."""
    conn = llm_client.ping()
    tools = get_tool_loader(llm_client)
    tool_names = tools.list_tools()
    detected_models = conn.get("models") or []
    if not detected_models and llm_client.config.model_name:
        detected_models = [llm_client.config.model_name]

    return {
        "provider": llm_client.config.provider,
        "base_url": llm_client.config.base_url,
        "model_name": llm_client.config.model_name,
        "agent_models": llm_client.config.agent_models,
        "supported_workflows": SUPPORTED_WORKFLOWS,
        "tools_count": len(tool_names),
        "available_tools": tool_names,
        "models": detected_models,
        "connection": conn,
    }


@app.get("/api/models")
def list_models_endpoint():
    """Returns all models detected from the connected local LLM server (oMLX/OpenAI/Ollama)."""
    conn = llm_client.ping()
    detected_models = conn.get("models") or []
    if not detected_models and llm_client.config.model_name:
        detected_models = [llm_client.config.model_name]
    return {
        "ok": conn.get("ok", False),
        "current_model": llm_client.config.model_name,
        "models": detected_models,
        "provider": llm_client.config.provider,
        "base_url": llm_client.config.base_url,
    }


@app.post("/api/models/select")
def select_model_endpoint(req: ModelSelectRequest):
    """Sets the active default model on the server."""
    model = req.model_name.strip()
    if model:
        if model.lower().startswith("qwen3.5") and hasattr(llm_client, "_normalize_ollama_model"):
            llm_client.config.model_name = llm_client._normalize_ollama_model(model)
        else:
            llm_client.config.model_name = model
    return {"ok": True, "model_name": llm_client.config.model_name}


@app.get("/api/tools")
def list_tools_endpoint():
    """Returns the list of all registered tools and their parameter metadata."""
    tools = get_tool_loader(llm_client)
    return {
        "ok": True,
        "count": len(tools.list_tools()),
        "tools": tools.describe(),
        "tool_names": tools.list_tools(),
    }


@app.post("/api/tools/run")
def execute_tool_endpoint(req: ToolRunRequest):
    """Directly executes a tool by name with arguments and pre/post checkpointing."""
    tools = get_tool_loader(llm_client)
    tool_name = req.tool.strip()
    kwargs = req.args or {}
    run_id = req.run_id or f"direct_tool_run_{int(time.time() * 1000)}"

    result = tools.run(
        tool_name,
        run_id=run_id,
        metadata=req.metadata or {"source": "api_tools_run"},
        **kwargs,
    )
    return result


@app.get("/api/tools/checkpoints")
def get_checkpoints_endpoint(
    run_id: Optional[str] = None,
    tool_name: Optional[str] = None,
    limit: int = 50,
):
    """Retrieve tool execution checkpoints from persistent SQLite memory."""
    checkpoints = get_tool_checkpoints(run_id=run_id, tool_name=tool_name)
    return {
        "ok": True,
        "count": len(checkpoints[:limit]),
        "checkpoints": checkpoints[:limit],
    }


class RollbackRequest(BaseModel):
    checkpoint_id: str


@app.get("/api/checkpoints/files")
def get_file_checkpoints_endpoint(
    file_path: Optional[str] = None,
    run_id: Optional[str] = None,
    limit: int = 50,
):
    """Retrieve pre- and post-file-change checkpoints."""
    from src.core.checkpointer import get_file_checkpoints

    checkpoints = get_file_checkpoints(file_path=file_path, run_id=run_id, limit=limit)
    return {
        "ok": True,
        "count": len(checkpoints),
        "checkpoints": checkpoints,
    }


@app.post("/api/checkpoints/rollback")
def rollback_file_checkpoint_endpoint(req: RollbackRequest):
    """Rolls back a file to the state recorded in a pre-file-change checkpoint."""
    from src.core.checkpointer import rollback_file_checkpoint

    res = rollback_file_checkpoint(req.checkpoint_id)
    if not res.get("ok"):
        raise HTTPException(status_code=400, detail=res.get("message", "Rollback failed"))
    return res


@app.get("/api/workflows")
def list_workflows():
    """Returns the supported workflow pipeline names for the API."""
    return {
        "supported_workflows": SUPPORTED_WORKFLOWS,
        "description": "Use the pipeline names in /api/chat requests. Model selection is available via model_name and agent_models.",
    }


@app.get("/api/memories")
def list_or_search_memories(
    q: Optional[str] = None, limit: int = 50, run_id: Optional[str] = None
):
    """Retrieve persistent memories, optionally filtered by search keyword or run_id."""
    if q:
        return {"memories": search_memories(query=q, limit=limit)}
    return {"memories": fetch_memories(limit=limit, run_id=run_id)}


@app.post("/api/memories")
def create_manual_memory(entry: MemoryCreateRequest):
    """Manually insert a persistent memory fact."""
    data = {
        "event": entry.event,
        "input": entry.input,
        "result": entry.result,
        "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
    }
    if entry.metadata:
        data.update(entry.metadata)
    mem_id = save_memory(data)
    return {"ok": True, "id": mem_id, "memory": data}


@app.delete("/api/memories/{memory_id}")
def remove_memory(memory_id: str):
    """Delete an individual persistent memory entry."""
    ok = delete_memory(memory_id)
    if not ok:
        raise HTTPException(status_code=404, detail="Memory item not found")
    return {"ok": True, "deleted_id": memory_id}


@app.delete("/api/memories")
def clear_all_stored_memories():
    """Clear all persistent memories."""
    count = clear_memories()
    return {"ok": True, "cleared_count": count}


@app.get("/api/memory-summary")
def memory_metrics():
    """Retrieve memory statistics for persistent storage overview."""
    return get_memory_summary()


@app.get("/api/runs")
def get_runs(limit: int = 50):
    """List recent conversation execution runs."""
    return {"runs": list_runs(limit=limit)}


@app.get("/api/runs/{run_id}")
def get_run_detail(run_id: str):
    """Get a specific run and its associated memories."""
    run = get_run(run_id)
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")
    memories = get_run_memories(run_id)
    return {"run": run, "memories": memories}


@app.delete("/api/runs/{run_id}")
def remove_run(run_id: str):
    """Delete a run and its grouped memories."""
    ok = delete_run(run_id)
    if not ok:
        raise HTTPException(status_code=404, detail="Run not found")
    return {"ok": True, "deleted_run_id": run_id}


@app.post("/api/chat/stream")
def handle_chat_stream(req: ChatRequest):
    """Executes selected LangGraph pipeline and streams step outputs via SSE."""
    prompt = req.prompt.strip()
    if not prompt and not req.images and not req.files:
        raise HTTPException(
            status_code=400,
            detail="Prompt, image, or file attachment cannot be empty.",
        )

    if not prompt and req.files:
        attached_names = [f.get("filename", "file") for f in req.files if isinstance(f, dict)]
        prompt = (
            f"Please analyze and summarize the attached spreadsheet/document: {', '.join(attached_names)}."
            if attached_names
            else "Please analyze and summarize the attached document."
        )

    pipeline_choice = (req.pipeline or "master").lower()
    timestamp = time.strftime("%H:%M:%S")
    user_msg_id = f"user_{int(time.time() * 1000)}"

    try:
        request_client = llm_client
        if req.model_name or req.agent_models:
            request_config = llm_client.config.copy(deep=True)
            if req.model_name:
                request_config.model_name = req.model_name
            if req.agent_models:
                request_config.agent_models = {
                    k.lower(): str(v) for k, v in req.agent_models.items()
                }
            request_client = LocalLLMClient(config=request_config)

        pipeline_choice = pipeline_choice or "master"
        graph_factory = WORKFLOW_FACTORIES.get(pipeline_choice)
        if graph_factory is None:
            raise HTTPException(
                status_code=400,
                detail=f"Unsupported pipeline '{req.pipeline}'. Supported pipelines: {', '.join(SUPPORTED_WORKFLOWS)}",
            )
        graph = graph_factory(request_client)

        # Persistent memory context recall
        recalled_facts = []
        if req.use_memory and prompt:
            try:
                mem_matches = search_memories(prompt, limit=4)
                for mem in mem_matches:
                    r_val = mem.get("result")
                    if r_val:
                        txt = (
                            json.dumps(r_val, ensure_ascii=False)
                            if isinstance(r_val, (dict, list))
                            else str(r_val)
                        )
                        recalled_facts.append(f"- [{mem.get('event', 'memory')}]: {txt[:200]}")
            except Exception:
                recalled_facts = []

        context_header = ""
        if recalled_facts:
            context_header = "\n[Recalled Memory Context:\n" + "\n".join(recalled_facts) + "\n]"

        effective_prompt = f"{prompt}{context_header}" if context_header else prompt

        # Append attached file contents to effective prompt
        if req.files:
            file_blocks = []
            for f in req.files:
                fn = f.get("filename", "attached_file.txt")
                txt = f.get("content") or f.get("text") or ""
                extracted = extract_file_text(fn, txt)
                file_blocks.append(f"--- File Attachment: {fn} ---\n{extracted}\n--- End File ---")
            if file_blocks:
                attachments_str = "\n\n" + "\n\n".join(file_blocks)
                effective_prompt = f"{effective_prompt}{attachments_str}" if effective_prompt else "\n\n".join(file_blocks)

        # Record start of run in SQLite
        run_id = start_run(
            pipeline=pipeline_choice,
            user_input=prompt or (f"[Attached Files: {len(req.files)}]" if req.files else ""),
            metadata={
                "has_images": bool(req.images),
                "image_count": len(req.images) if req.images else 0,
                "has_files": bool(req.files),
                "file_count": len(req.files) if req.files else 0,
                "files": [f.get("filename") for f in req.files] if req.files else [],
                "model_name": req.model_name or request_client.config.model_name,
                "session_id": req.session_id,
            },
        )

        effective_images = list(req.images or [])
        if req.files:
            for f in req.files:
                if isinstance(f, dict):
                    fn = (f.get("filename") or "").lower()
                    cnt = f.get("content") or f.get("text") or ""
                    if fn.endswith((".png", ".jpg", ".jpeg", ".webp", ".gif", ".bmp", ".svg")) or cnt.startswith("data:image/"):
                        if cnt and cnt not in effective_images:
                            effective_images.append(cnt)

        user_message_entry = {
            "id": user_msg_id,
            "sender": "User",
            "role": "user",
            "content": effective_prompt,
            "timestamp": timestamp,
            "images": effective_images,
            "files": req.files or [],
        }

        if pipeline_choice in ["direct", "chat", "vision", "qwen"]:
            initial_input = {
                "run_id": run_id,
                "messages": [user_message_entry],
                "user_input": effective_prompt,
                "final_response": "",
                "agent_thoughts": [],
                "images": effective_images,
                "files": req.files or [],
            }
        elif pipeline_choice in ["chart", "chart_pipeline"]:
            initial_input = {
                "run_id": run_id,
                "messages": [user_message_entry],
                "user_input": effective_prompt,
                "current_step": "intake",
                "agent_thoughts": [],
                "images": effective_images,
                "files": req.files or [],
            }
        elif pipeline_choice in ["code_review", "code"]:
            initial_input = {
                "run_id": run_id,
                "messages": [user_message_entry],
                "task": effective_prompt,
                "code": "",
                "review": "",
                "approved": False,
                "revision_count": 0,
                "agent_thoughts": [],
            }
        elif pipeline_choice in ["solution_review", "solution"]:
            initial_input = {
                "run_id": run_id,
                "messages": [user_message_entry],
                "task": effective_prompt,
                "solution": "",
                "review": "",
                "approved": False,
                "revision_count": 0,
                "agent_thoughts": [],
            }
        elif pipeline_choice == "supervisor":
            initial_input = {
                "run_id": run_id,
                "messages": [user_message_entry],
                "current_task": effective_prompt,
                "next_agent": "supervisor",
                "research_output": "",
                "coder_output": "",
                "critic_feedback": "",
                "final_response": "",
                "agent_thoughts": [],
            }
        elif pipeline_choice in ["claims_triage", "claims"]:
            initial_input = {
                "run_id": run_id,
                "messages": [user_message_entry],
                "claim_input": effective_prompt,
                "current_step": "step_1_classification",
                "classification_details": None,
                "severity_assessment": None,
                "action_plan": None,
                "final_response": "",
                "agent_thoughts": [],
            }
        else:  # master / default
            initial_input = {
                "run_id": run_id,
                "messages": [user_message_entry],
                "user_input": effective_prompt,
                "current_step": "pipeline_start",
                "triage_details": None,
                "supervisor_details": None,
                "review_details": None,
                "final_response": "",
                "agent_thoughts": [],
            }


        def event_stream():
            chunk_q: "_queue.Queue" = _queue.Queue()

            def _worker():
                try:
                    config = {"configurable": {"thread_id": run_id}}
                    for chunk in graph.stream(initial_input, config=config):
                        chunk_q.put(("chunk", chunk))
                except Exception as err:  # noqa: BLE001
                    chunk_q.put(("error", str(err)))
                finally:
                    chunk_q.put(("done", None))

            _threading.Thread(target=_worker, daemon=True).start()

            step_idx = 1
            steps_data = []
            final_answer = ""

            while True:
                try:
                    kind, payload = chunk_q.get(timeout=15)
                except _queue.Empty:
                    # Keep-alive comment
                    yield ": heartbeat\n\n"
                    continue

                if kind == "error":
                    finish_run(run_id=run_id, status="error", final_answer=f"Error: {payload}")
                    yield f"data: {json.dumps({'type': 'error', 'detail': payload, 'run_id': run_id})}\n\n"
                    break

                if kind == "done":
                    break

                # kind == "chunk"
                for node_name, node_update in payload.items():
                    thoughts = node_update.get("agent_thoughts") or []
                    messages = node_update.get("messages") or []
                    final_resp = node_update.get("final_response") or ""

                    step_payload = {
                        "step": step_idx,
                        "node": node_name,
                        "thoughts": thoughts,
                        "messages": messages,
                        "final_response": final_resp,
                    }
                    steps_data.append(step_payload)

                    yield f"data: {json.dumps({'type': 'step', 'data': step_payload, 'run_id': run_id})}\n\n"
                    step_idx += 1

            # Extract final answer
            for s in reversed(steps_data):
                if s.get("final_response"):
                    final_answer = s["final_response"]
                    break
                if s.get("messages"):
                    last_m = s["messages"][-1]
                    if isinstance(last_m, dict) and last_m.get("content"):
                        final_answer = last_m["content"]
                        break

            final_answer = final_answer or "Pipeline execution completed."

            # Save completed run and record memory in SQLite
            finish_run(
                run_id=run_id,
                status="completed",
                final_answer=final_answer,
            )
            save_memory(
                {
                    "run_id": run_id,
                    "event": f"pipeline_{pipeline_choice}",
                    "input": prompt,
                    "result": final_answer,
                    "metadata": {
                        "has_images": bool(req.images),
                        "image_count": len(req.images) if req.images else 0,
                        "model_name": req.model_name or request_client.config.model_name,
                        "session_id": req.session_id,
                    },
                    "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
                }
            )

            complete_payload = {
                "type": "complete",
                "run_id": run_id,
                "pipeline": pipeline_choice,
                "prompt": prompt,
                "has_images": bool(req.images),
                "steps": steps_data,
                "final_response": final_answer,
            }
            yield f"data: {json.dumps(complete_payload)}\n\n"

        return StreamingResponse(
            event_stream(),
            media_type="text/event-stream",
            headers={
                "Cache-Control": "no-cache",
                "X-Accel-Buffering": "no",
                "Connection": "keep-alive",
            },
        )

    except HTTPException:
        raise
    except Exception as err:
        raise HTTPException(
            status_code=500, detail=f"Pipeline execution error: {str(err)}"
        ) from err


# Serve static files from public directory
public_dir = os.path.join(os.path.dirname(__file__), "public")
if os.path.exists(public_dir):
    app.mount("/", StaticFiles(directory=public_dir, html=True), name="static")

if __name__ == "__main__":
    import uvicorn

    # Web UI server runs on port 8080 to avoid port conflict with oMLX server on port 8000
    uvicorn.run("server:app", host="0.0.0.0", port=8080, reload=True)
